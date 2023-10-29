import pandas as pd
import numpy as np
from openpyxl import load_workbook

workbook = load_workbook('excel.xlsx')
sheet = workbook.active

for row in sheet.iter_rows(values_only=True):
    for cell in row:
        print(cell, end='\t')
    print()